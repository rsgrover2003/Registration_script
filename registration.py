# Copyright (c) 2016 Radhika S. Grover, Santa Clara University.
# Registration script for IEEE RFID conference
# See the file IEEE_RFID_Functional_Specifications for details
# Takes the registration list, registration rates, accepted papers and posters as input
# Generates a list of covered and non-covered papers and posters
# Uses the openpyxl and itertools packages
# Python version 3.5
# Important formatting for dates. In Google docs in the Excel file convert date as follows: Format->Number->Date m/d/yyyy.
# Then convert date to text using Format->Number->Plain text. The date should be in plain text to work correctly
# with this script

# Assumption: max of 10 authors per paper or poster
# Assumption: last name + first name is unique
# Check that first and last names with unicode characters are written with unicode and not plain text in other files.
# script ignores middle names and affiliations


from openpyxl import load_workbook
from datetime import datetime
from collections import Counter
import itertools

wb = load_workbook(filename = '/Users/rgrover/Documents/ieeerfidexcelfiles/IEEE_RFID_Registration_Rates.xlsx')
reg_rates = wb.active
wb1 = load_workbook(filename = '/Users/rgrover/Documents/ieeerfidexcelfiles/IEEE_RFID_Papers.xlsx')
reg_papers = wb1.active
wb2 = load_workbook(filename = '/Users/rgrover/Documents/ieeerfidexcelfiles/IEEE_RFID_Posters.xlsx')
reg_posters = wb2.active
wb3 = load_workbook(filename = '/Users/rgrover/Documents/ieeerfidexcelfiles/LIEEE_RFID_Registration_List.xlsx')
reg_list = wb3.active

MAX_AUTHORS = 10  #max of 10 authors per paper or poster

i = 0
for cell in reg_rates.rows[0]:
    if cell.value == 'Non-member':
        print(i)
        c = reg_rates.columns[i]
        #for cell in c:
          #print(cell.value)
    i = i+1

j = 0
for cell in reg_papers.rows[0]:
    if cell.value == 'Author 1':
        c = reg_papers.columns[j]
        #for cell in c:
          #print(cell.value)
    j = j+1

k = 0
for cell in reg_posters.rows[0]:
    if cell.value == 'Author 1':
        print(k)
        c = reg_posters.columns[k]
        #for cell in c:
          #print(cell.value)
    k = k+1


l = 0
# read the registration date (RegDate column) into the array regdates_list
regdates_list = []
for cell in reg_list.rows[0]:
    if cell.value == 'RegDate':
        print(l)

        dates = reg_list.columns[l]
        for cell in dates:
            # parse textual date value to create datetime object as yyyy-mm-dd
            if cell.value == 'RegDate':
                print(' ')
            else:
                strval = cell.value
                if strval != None:
                #print("Here" +strval)
                    arr1 = strval.split('/')
                    arr2 = arr1[2].split()

                    date_object = datetime(int(arr2[0]), int(arr1[0]), int(arr1[1]))
                    regdates_list.append(date_object)

    l = l+1

#for data in regdates_list:
    #print (data)
first_names_list = []  #names stored in rows 0,...
last_names_list = []    #names stored in rows 0,...
m = 0
for cell in reg_list.rows[0]:
    if cell.value == 'First Name':
        names = reg_list.columns[m]
        for cell in names:
            if cell.value != 'First Name' and cell.value != 'None':
                first_names_list.append(cell.value)
    elif cell.value == 'Last Name':
        names = reg_list.columns[m]
        for cell in names:
            if cell.value != 'Last Name' and cell.value != 'None':
                last_names_list.append(cell.value)


    m = m+1


# read the total amount paid for the registration (column Total1) into the array amount_paid_list
l = 0
amount_paid_list = []
for cell in reg_list.rows[0]:
    if cell.value == 'Total1':
        print(l)

        dates = reg_list.columns[l]
        for cell in dates:
            # parse textual date value to create datetime object as yyyy-mm-dd
            if cell.value == 'Total1':
                print(' ')
            else:
                strval = cell.value
                if strval != None:
                    #print(strval)
                    amount_paid_list.append(strval)
    l = l+1

#for data in amount_paid_list:
    #print (data)

l = 0
# read the registration rates (all columns in Registration_rates file) into 2D array reg_rates_list
reg_rates_list = []
for cell in reg_rates.columns[0]:
    reg_rates_list.append(reg_rates.rows[l])
    l = l+1
# store the dates in column "End date" as datetime objects
m = 0
reg_rates_date_list = []
for cell in reg_rates.rows[0]:
    if cell.value == 'End date':
        dates = reg_rates.columns[m]
        for cell in dates:
            # parse textual date value to create datetime object as yyyy-mm-dd
            if cell.value == 'End date':
                print(' ')
            else:
                strval = cell.value
                if strval != None:
                #print("Here" +strval)
                    arr1 = strval.split('/')


                    date_object = datetime(int(arr1[2]), int(arr1[0]), int(arr1[1]))
                    reg_rates_date_list.append(date_object)

    m = m+1
#for data in reg_rates_date_list:
    #print(data)

# read all data in papers (all columns in IEEE_RFID_Papers file) into 2D array called papers_list
l = 0
papers_list = []
for cell in reg_papers.columns[0]:
    papers_list.append(reg_papers.rows[l])
    l = l+1

# read all data in posters (all columns in IEEE_RFID_Posters file) into 2D array called posters_list
l = 0
posters_list = []
for cell in reg_posters.columns[0]:
    posters_list.append(reg_posters.rows[l])
    l = l+1

p = 0
paper_and_poster_ids = []   # this array stores all of the paper and poster ids
#create an array containing all the paper ids
for cell in reg_papers.rows[0]:
    if cell.value == 'ID':
        ids = reg_papers.columns[p]
        for cell in ids:
            # parse textual date value to create datetime object as yyyy-mm-dd
            if cell.value == 'ID':
                print(' ')
            else:
                paper_and_poster_ids.append(cell.value)

    p = p+1

p = 0
#add all poster ids to array paper_and_poster_ids
for cell in reg_posters.rows[0]:
     if cell.value == 'ID':
        ids = reg_posters.columns[p]
        for cell in ids:
            # parse textual date value to create datetime object as yyyy-mm-dd
            if cell.value == 'ID':
                print(' ')
            else:
                paper_and_poster_ids.append(cell.value)

     p = p+1

# print out all the paper and poster ids
for data in paper_and_poster_ids:
    print (data)

print('End')

# this function takes the member_status and the reg_date
def get_registration_rates(member_status, reg_date):
    print('The member status is' +member_status)
    if "Non-member".lower() in member_status.lower():
        col = 4
    elif "L/S/R".lower() in member_status.lower():
        col = 3
    elif "Member".lower() in member_status.lower():
        col = 2
    elif "All-Access".lower() in member_status.lower():
        col = 5
    else:
        raise ValueError("Incorrect member status")

    if reg_date <= reg_rates_date_list[0]:
        row = 1   #Early registration
    elif reg_date <= reg_rates_date_list[1]:
        row = 2    #Advanced registration
    elif reg_date <= reg_rates_date_list[2]:
        row = 3    #Regular
    else:
        row = 4    #On-site

    # return the registration fee for given row and column
    return reg_rates_list[row][col].value

# find number of registrations for each member and store in the array number_of_registrations
number_of_registrations = []
l = 0
# read the member status (Description1 column) into the array member_status_list
member_status_list = []
done = False
for cell in reg_list.rows[0]:
    if cell.value == 'Description1' and done == False:
        print(l)
        done = True
        status = reg_list.columns[l]
        k = 0
        for cell in status:
            if cell.value == 'Description1':
                print(' ')
            elif cell.value != None:
                print(done)
                print('index'+str(k))
                print(cell.value)
                print(regdates_list[k-1])
                # L/S/R does not count as registered
                if "L/S/R" in cell.value:
                    number_of_registrations.append(0)
                else:
                    # find the amount to be paid based on attendee's registration date
                    amount_to_pay = get_registration_rates(cell.value, regdates_list[k-1])
                    print(amount_to_pay)
                    # find the number of registrations for each attendee based on amount actually paid by attendee
                    # for example, with 2 registrations, an author can cover 4 papers
                    num_regs = int(amount_paid_list[k-1]/amount_to_pay)
                    print(num_regs)
                    number_of_registrations.append(num_regs)
            k = k+1

    l = l+1

for data in first_names_list:
    print (data)

for data in last_names_list:
    print (data)

index = 0
#find the number of papers+posters for each member



#add all the authors of papers to the array author_list
index = 0
author_list = []
for data in reg_papers.rows[0]:
    if data.value != None and "Author" in data.value:
        arr = reg_papers.columns[index]
        for cell in arr:
            cellval = cell.value
            if cellval == None or 'Author' in cellval:
                print('')
            else:
                author_list.append(str(cellval))   # add author name to author list
    index = index + 1



#add all the authors of posters to the array author_list
index = 0
for data in reg_posters.rows[0]:
    if data.value != None and "Author" in data.value:
        arr = reg_posters.columns[index]
        for cell in arr:
            cellval = cell.value
            if cellval == None or 'Author' in cellval:
                print('')
            else:
                author_list.append(str(cellval))
    index = index + 1


for data in author_list:
    print(data)

# add the tuples {author name: [list of paper ids]} to dictionary  author_papers_dict
# add the tuples {paper id: [list of author names'} to dictionary paper_authors_dict
index = 0
author_papers_dict = {}   # {author name: [list of paper ids]}
paper_authors_dict = {}   # {paper id: [list of authors]}
for data in reg_papers.rows[0]:
    if data.value != None and "ID" in data.value:
        ids = reg_papers.columns[index]
    if data.value != None and "Author" in data.value:
        arr = reg_papers.columns[index]
        row_num = 0
        for cell in arr:
            cellval = cell.value
            if cellval == None or 'Author' in cellval:
                print('')
            else:
                paper_num = ids[row_num].value
                # set key to author name and value to paper number
                author_papers_dict.setdefault(str(cellval), []).append(paper_num)
                 # set key to paper id and value to author name
                paper_authors_dict.setdefault(paper_num, []).append(str(cellval))
            row_num = row_num + 1
    index = index + 1

# add the tuples {author name: [list of poster ids]} to dictionary  author_papers_dict
# add the tuples {poster id: [list of author names'} to dictionary papers_author_dict
index = 0
for data in reg_posters.rows[0]:
    if data.value != None and "ID" in data.value:
        ids = reg_posters.columns[index]
    if data.value != None and "Author" in data.value:
        arr = reg_posters.columns[index]
        row_num = 0
        for cell in arr:
            cellval = cell.value
            if cellval == None or 'Author' in cellval:
                print('')
            else:
                poster_num = ids[row_num].value
                # set key to author name and value to poster number
                author_papers_dict.setdefault(str(cellval), []).append(poster_num)
                # set key to poster id and value to author name
                paper_authors_dict.setdefault(poster_num, []).append(str(cellval))
            row_num = row_num + 1
    index = index + 1

# show dictionary
print("Author name: list of papers and posters")
print(author_papers_dict.items())
print("\n \n")

# show dictionary
print("Paper id: list of authors")
print(paper_authors_dict.items())
print("\n \n")

# count the number of papers+posters for each author
# here the assumption is that a author is identified uniquely by first and last names
# for example, Max Rey has 2 posters and 2 papers; total_paper_and_posters = 4
count = Counter(author_list)
print("Author name: total number of papers and posters")
print(count)

index = 0
total_papers_and_posters = []   #total number of papers and posters for that author
# display number of registrations and total number of papers and posters for each member
for data in number_of_registrations:
    print("First name")
    print(first_names_list[index])
    print("Last name ")
    print(last_names_list[index])
    print("Number of registrations")
    print(data)
    first_name = first_names_list[index]
    last_name = last_names_list[index]
    full_name = str(first_names_list[index]) + " " + str(last_names_list[index])
    print(full_name)
    try:
        print("Total papers and posters")
        value = 0
        #value += next(v for k,v in count.items() if (k.startswith(first_name) and last_name in k))
        for k,v in count.items():
            # make a match when k="Ray Lay", and first_name = "Ray Kay" using the or clause
            if (( k.startswith(first_name) or k.startswith(first_name.partition(' ')[0])) and last_name in k):
                value = value + v
                print("value =" +str(value))
        total_papers_and_posters.append(value)
        print(total_papers_and_posters[-1])
    except StopIteration:
        print("Stop Iteration Exception")
        total_papers_and_posters.append(0)
    except TypeError:
        print("Found None type instead of str or tuple of str")
    print("index = " +str(index))
    index = index + 1



# for all papers/posters with 0 registered authors
# move papers and posters with 0 registered authors to non-covered list
# store ids of all uncovered papers in the array not_covered
not_covered = []
index = 0
for paperid in paper_authors_dict.keys():
    print("Paper/poster id :" + str(paperid))
    authors = paper_authors_dict.get(paperid)
    print (authors)
    any_registered = False
    for author in authors:
        print("Author is " +author)
        index = 0
        #for data in number_of_registrations:
        #if (author.lower()).startswith(first_names_list[index].lower()) and last_names_list[index].lower() in author.lower():
        # use lower() because one last name is spelled as XI in last_names_index[] but as Xi in author
        for firstName in first_names_list:
            if firstName != None:
                if (( (author.lower()).startswith(first_names_list[index].lower()) or (author.lower()).startswith((first_names_list[index].partition(' ')[0]).lower())) and last_names_list[index].lower() in author.lower()):
                    #print("index " +str(index))
                    print("number of registrations: " +str(number_of_registrations[index]))
                    if number_of_registrations[index] > 0:  # registered author
                        any_registered = True

            index = index + 1

    if any_registered == False:
        print(str(paperid) +str("does not have registered authors"))
        # this paper/poster does not have any registered authors - add to not_covered list
        not_covered.append(paperid)


print("IDs of papers and posters that are not covered")
for data in not_covered:
    print (data)
    #remove this paper/poster from the paper_authors_dict
    paper_authors_dict.pop(data, None)

# for each member with 2 * number of registrations >= number of papers & posters
# assign all papers/posters for that member as covered
index = 0
covered = []   # contains IDs of covered papers and posters
for data in number_of_registrations:
    print("First name")
    print(first_names_list[index])
    print("Last name ")
    print(last_names_list[index])
    print("Number of registrations")
    print(data)
    first_name = first_names_list[index]
    last_name = last_names_list[index]
    print("Total number of papers and posters")
    print(total_papers_and_posters[index])
    try:
        if (data * 2) >= total_papers_and_posters[index]:
            # move the papers/posters for this author to covered list
            for k,v in author_papers_dict.items():
                # make a match when k="Ray Lay", and first_name = "Ray Kay" using the or clause
                if (( (k.lower()).startswith(first_name.lower()) or (k.lower()).startswith((first_name.partition(' ')[0]).lower())) and last_name.lower() in k.lower()):
                    # add this paper id to covered list
                    covered.append(v)
                    # remove this paper id from paper_authors_dict
                    #paper_authors_dict.pop(paperid, None)
                    #covered.append(next(v for k,v in author_papers_dict.items() if (k.startswith(first_name) and last_name in k)))
                    print (covered[-1])
                    # change the number of registrations for this author to 0  probably don't need this
                    # CORRECT THIS IN COUNT
                    number_of_registrations[index] = 0
    except StopIteration:
        print("Stop Iteration Exception")
    print ("index =" +str(index))
    index = index + 1

print("IDs of covered papers and posters")
for data in covered:
    print (data)
    for element in data:
        paper_authors_dict.pop(element, None)

print("IDs of papers with unknown status")
for data in paper_authors_dict:
    print(data)

# This section is to be completed to further categorize papers with "unknown" status into one of the "covered" or "not-covered" lists.
# Use lower() to match first and last names
# use the permutations method in python package itertools to get best assignment for the rest
list1 = [52, 119, 125]
list2 = ['JE', 'JE', 'DP', 'DP', 'GD', 'GD', 'FA', 'FA']
value = len(list1)
#create author tuples
array2 = itertools.combinations(list2, value)

count = 0
best = 0 # best tuple
#for i in array2:
    # read each author in each tuple
    #for j in range(0, value):
        #author = str(i[j])
        # if this author is an author for list1[j] increment count
        #authors = paper_authors_dict.get(paperid)
        #print (authors)
        #any_registered = False
        #for author in authors:
            #index = 0
            # use lower() because one last name is spelled as XI in last_names_index[] but as Xi in author
            #for firstName in first_names_list:
                #if firstName != None:
                    #if (( (author.lower()).startswith(first_names_list[index].lower()) or (author.lower()).startswith((first_names_list[index].partition(' ')[0]).lower())) and last_names_list[index].lower() in author.lower()):
                        #print("index " +str(index))
                        #print("number of registrations: " +str(number_of_registrations[index]))
                        #if number_of_registrations[index] > 0:  # registered author
                            #any_registered = True

            #index = index + 1

    # if count > best, update best to store this tuple
    # complete!



